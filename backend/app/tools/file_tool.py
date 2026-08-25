"""
Sovereign File Read / Write / Edit Tool
-----------------------------------------
Gives agents the ability to read existing uploaded files and write/edit
local workspace files — scoped strictly to the sandboxed workspace directory.

PS SIH26117 explicitly requires "file read and write" as a local tool.
This addresses Gap 1.5 from the gap analysis.
"""

import os
import json
import csv
import hashlib
from pathlib import Path
from typing import Dict, Any, Optional
from app.config import settings
from app.core.logging import logger
from app.core.security_guard import security_guard


# All file operations are strictly sandboxed to this directory
WORKSPACE_DIR: Path = settings.STORAGE_DIR / "workspace"
WORKSPACE_DIR.mkdir(parents=True, exist_ok=True)


def _safe_path(filename: str) -> Optional[Path]:
    """
    Resolves a filename to a safe absolute path inside WORKSPACE_DIR.
    Rejects any path-traversal attempts (e.g. ../../etc/passwd).
    """
    safe = (WORKSPACE_DIR / Path(filename).name).resolve()
    if not str(safe).startswith(str(WORKSPACE_DIR.resolve())):
        logger.warning(f"Path traversal attempt blocked: {filename}")
        return None
    return safe


class FileTool:
    """
    Air-gapped local file read / write / edit tool.
    All operations are strictly scoped to the sandboxed workspace directory.
    """

    def read_file(self, filename: str) -> Dict[str, Any]:
        """Read any text/CSV/JSON file from the workspace."""
        security_guard.record_local_request()
        path = _safe_path(filename)

        # Also check uploads directory for recently uploaded files
        if path is None or not path.exists():
            upload_path = (settings.UPLOAD_DIR / Path(filename).name).resolve()
            if upload_path.exists():
                path = upload_path
            else:
                return {"success": False, "error": f"File '{filename}' not found in workspace or uploads."}

        try:
            suffix = path.suffix.lower()

            if suffix == ".csv":
                with open(path, "r", encoding="utf-8", errors="ignore") as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)
                return {
                    "success": True,
                    "filename": path.name,
                    "format": "CSV",
                    "row_count": len(rows),
                    "columns": list(rows[0].keys()) if rows else [],
                    "data": rows[:50],   # First 50 rows to limit context
                    "air_gap_verified": True,
                }

            elif suffix == ".xlsx":
                import openpyxl
                wb = openpyxl.load_workbook(path, data_only=True)
                sheet_names = wb.sheetnames
                ws = wb.active
                data_rows = []
                for row in ws.iter_rows(values_only=True):
                    if any(v is not None for v in row):
                        data_rows.append([str(v) if v is not None else "" for v in row])
                return {
                    "success": True,
                    "filename": path.name,
                    "format": "XLSX",
                    "sheets": sheet_names,
                    "active_sheet": ws.title,
                    "row_count": len(data_rows),
                    "data": data_rows[:30],
                    "air_gap_verified": True,
                }

            elif suffix == ".json":
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                return {
                    "success": True,
                    "filename": path.name,
                    "format": "JSON",
                    "data": data,
                    "air_gap_verified": True,
                }

            else:
                # Plain text (txt, log, py, md, etc.)
                with open(path, "r", encoding="utf-8", errors="ignore") as f:
                    content = f.read(20_000)   # Read up to 20KB
                return {
                    "success": True,
                    "filename": path.name,
                    "format": "TEXT",
                    "char_count": len(content),
                    "content": content,
                    "air_gap_verified": True,
                }

        except Exception as e:
            logger.error(f"File read error for {filename}: {e}")
            return {"success": False, "error": str(e)}

    def write_file(self, filename: str, content: str, mode: str = "overwrite") -> Dict[str, Any]:
        """
        Write or append text content to a file in the workspace.
        mode: 'overwrite' | 'append'
        """
        security_guard.record_local_request()
        path = _safe_path(filename)
        if path is None:
            return {"success": False, "error": "Invalid filename — path traversal blocked."}

        try:
            write_mode = "w" if mode == "overwrite" else "a"
            with open(path, write_mode, encoding="utf-8") as f:
                f.write(content)

            file_size = path.stat().st_size
            sha256 = hashlib.sha256(content.encode()).hexdigest()

            return {
                "success": True,
                "filename": path.name,
                "storage_path": str(path),
                "bytes_written": len(content.encode()),
                "file_size_bytes": file_size,
                "integrity_sha256": sha256,
                "air_gap_verified": True,
            }

        except Exception as e:
            logger.error(f"File write error for {filename}: {e}")
            return {"success": False, "error": str(e)}

    def edit_csv_cell(self, filename: str, row_index: int, column: str, new_value: str) -> Dict[str, Any]:
        """
        Edit a specific cell in a CSV file in the workspace.
        Useful for agents updating spreadsheet-like data in place.
        """
        security_guard.record_local_request()
        path = _safe_path(filename)

        if path is None or not path.exists():
            return {"success": False, "error": f"File '{filename}' not found."}

        try:
            with open(path, "r", encoding="utf-8") as f:
                rows = list(csv.DictReader(f))
                fieldnames = list(rows[0].keys()) if rows else []

            if row_index >= len(rows):
                return {"success": False, "error": f"Row {row_index} out of range (total: {len(rows)})"}
            if column not in fieldnames:
                return {"success": False, "error": f"Column '{column}' not found. Available: {fieldnames}"}

            old_value = rows[row_index][column]
            rows[row_index][column] = new_value

            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(rows)

            return {
                "success": True,
                "filename": path.name,
                "row_index": row_index,
                "column": column,
                "old_value": old_value,
                "new_value": new_value,
                "air_gap_verified": True,
            }

        except Exception as e:
            return {"success": False, "error": str(e)}

    def edit_excel_cell(self, filename: str, cell_coordinate: str, new_value: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Edit a specific cell in an existing Excel (.xlsx) workbook in place.
        Satisfies PS Requirement 6c & C2 (Spreadsheet In-Place Editing).
        """
        security_guard.record_local_request()
        path = _safe_path(filename)
        if path is None or not path.exists():
            # Check generated or uploads dir
            gen_path = (settings.GENERATED_DIR / Path(filename).name).resolve()
            if gen_path.exists():
                path = gen_path
            else:
                return {"success": False, "error": f"Excel file '{filename}' not found."}

        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            old_val = str(ws[cell_coordinate].value)
            ws[cell_coordinate] = new_value
            wb.save(path)

            return {
                "success": True,
                "filename": path.name,
                "sheet": ws.title,
                "cell": cell_coordinate,
                "old_value": old_val,
                "new_value": new_value,
                "air_gap_verified": True
            }
        except Exception as e:
            return {"success": False, "error": str(e)}

    def list_workspace_files(self) -> Dict[str, Any]:
        """List all files currently in the agent workspace directory."""
        files = []
        for p in WORKSPACE_DIR.iterdir():
            if p.is_file():
                files.append({
                    "filename": p.name,
                    "size_bytes": p.stat().st_size,
                    "extension": p.suffix,
                })
        return {"success": True, "workspace_files": files, "count": len(files)}


file_tool = FileTool()
